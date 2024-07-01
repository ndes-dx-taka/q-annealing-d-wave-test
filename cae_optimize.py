from collections import defaultdict
from dwave.system.samplers import DWaveSampler
from dwave.system.composites import EmbeddingComposite
from dwave.system import LeapHybridSampler
import logging
import matplotlib.pyplot as plt
import math
import numpy as np
import openpyxl
from openpyxl.drawing.image import Image
import os
import pandas as pd
import pyautogui
import re
from shapely.geometry import Point, Polygon
import statistics
import subprocess
import sys
import shutil
import time
import win32con
import win32gui
from xml.etree import ElementTree as ET

initial_condition_data = {
    "width": 70,
    "height": 40,
    "target_density": 0.5,
    "density_increment": 0.1,
    "density_power": 2.0,
    "initial_youngs_modulus": 2.0e+5,
    "initial_volume": 2800.0,
    "cost_lambda": 5,
    "cost_lambda_n": 100,
    "loop_num": 20,
    "decide_val_threshold": 0.1,
    "start_phase_num": 1,
    "alwaysUpdateExcel" : 1,
    "devide_num_x" : 700,
    "devide_num_y" : 400
}

logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler("C:\\work\\github\\q-annealing-d-wave-test\\cae_opti_info.log"),
                        logging.StreamHandler()
                    ])

def check_input_excel(sheet):
    start_row = 20
    for row in range(start_row + 1, sheet.max_row + 1):
        a_value = sheet.cell(row=row, column=1).value
        if str(a_value) != str(row - start_row):
            return False
    return True

def find_elems_in_elset(elset_element, elem_list):
    for child in elset_element:
        if child.tag == 'elset':
            find_elems_in_elset(child, elem_list)
        elif child.tag == 'elem':
            elem_info = {
                'eid': child.attrib.get('eid', None),
                'nodes': [int(n) for n in child.attrib.get('nodes', '').split()],
                'material': elset_element.attrib.get('material', None)
            }
            elem_list.append(elem_info)

def find_mats_in_solution(solution_element, mat_list):
    for mat in solution_element.findall(".//mat"):
        mat_info = {'name': mat.attrib.get('name', None)}
        geometric = mat.find('geometric')
        if geometric is not None:
            mat_info.update({
                'geometric_type': geometric.attrib.get('type', None),
                'thickness': geometric.attrib.get('thickness', None),
                'planestrain': geometric.attrib.get('planestrain', None)
            })
        mechanical = mat.find('mechanical')
        if mechanical is not None:
            mat_info.update({
                'mechanical_type': mechanical.attrib.get('type', None),
                'youngsmodulus': mechanical.attrib.get('youngsmodulus', None),
                'poissonratio': mechanical.attrib.get('poissonratio', None)
            })
        mat_list.append(mat_info)

def calculate_quadrilateral_area(vertices):
    x_coords = [vertex['x'] for vertex in vertices]
    y_coords = [vertex['y'] for vertex in vertices]
    area = 0.5 * abs(sum(x1 * y2 - x2 * y1 for x1, y1, x2, y2 in zip(x_coords, y_coords, x_coords[1:] + x_coords[:1], y_coords[1:] + y_coords[:1])))
    return area

def increment_phase_number(filename):
    match = re.search(r'phase_(\d+)\.liml', filename)
    
    new_filename = ""
    if match:
        original_number = int(match.group(1))
        incremented_number = original_number + 1
        new_filename = filename.replace(f'phase_{original_number}', f'phase_{incremented_number}')
    else:
        new_filename = filename.replace('.liml', '_phase_1.liml')
    return new_filename

def renew_excel():
    # ファイルのパスを定義
    original_file = sys.argv[2]
    backup_file = "C:\\work\\github\\q-annealing-d-wave-test\\result_summary_bak.xlsx"

    # 元のファイルを削除
    if os.path.exists(original_file):
        os.remove(original_file)
        print(f"{original_file} を削除しました。")
    else:
        print(f"{original_file} は存在しません。")

    # バックアップファイルをコピーして新しい名前で保存
    if os.path.exists(backup_file):
        shutil.copy(backup_file, original_file)
        print(f"{backup_file} を {original_file} としてコピーしました。")
    else:
        print(f"{backup_file} は存在しません。")

def expected_value(n):
    sum_value = 0
    for k in range(n + 1):
        term = math.comb(n, k) * ((2 * k - n) ** 2) / (2 ** n)
        sum_value += term
    return sum_value

def expected_value_2_zyou(n):
    sum_value = 0
    for k in range(n + 1):
        term = math.comb(n, k) * ((2 * k - n) ** 4) / (2 ** n)
        sum_value += term
    return sum_value

def calc_seiyaku_bunsan(n):
    exp = expected_value(n)
    exp_2_zyou = expected_value_2_zyou(n)
    return (exp_2_zyou - exp)

def main(file_path):
    loop_num = initial_condition_data['loop_num']
    start_phase_num = initial_condition_data['start_phase_num'] - 1
    bAlwaysUpdateExcel = False
    bAlwaysUpdateExcelval = initial_condition_data['alwaysUpdateExcel']
    if str(bAlwaysUpdateExcelval) == "1":
        bAlwaysUpdateExcel = True
    if loop_num <= start_phase_num or bAlwaysUpdateExcel:
        renew_excel()
    # for i in range(loop_num):
    for i in range(start_phase_num, loop_num):
        result = main2(file_path, i)
        if result == False:
            logging.error(f"{i}/{loop_num}の処理で失敗しました")

def find_vscode_window():
    hwnd = None

    def callback(handle, extra):
        nonlocal hwnd
        if "Visual Studio Code" in win32gui.GetWindowText(handle):
            hwnd = handle

    win32gui.EnumWindows(callback, None)
    return hwnd

def main2(file_path, phase):
    start_time_1 = time.time()
    workbook = openpyxl.load_workbook(sys.argv[2])
    sheet = workbook["Sheet1"]

    target_density = initial_condition_data["target_density"]
    density_increment = initial_condition_data["density_increment"]
    density_power = initial_condition_data["density_power"]
    initial_youngs_modulus = initial_condition_data["initial_youngs_modulus"]
    initial_volume = initial_condition_data["initial_volume"]
    cost_lambda = initial_condition_data["cost_lambda"]
    cost_lambda_n = initial_condition_data["cost_lambda_n"]
    loop_num = initial_condition_data["loop_num"]
    decide_val_threshold = initial_condition_data["decide_val_threshold"]

    threshold = 0.001

    sheet.cell(row=5, column=1, value=f"{str(target_density)}")
    sheet.cell(row=5, column=2, value=f"{str(density_increment)}")
    sheet.cell(row=5, column=3, value=f"{str(density_power)}")
    sheet.cell(row=5, column=4, value=f"{str(initial_youngs_modulus)}")
    sheet.cell(row=5, column=5, value=f"{str(initial_volume)}")
    sheet.cell(row=5, column=6, value=f"{str(cost_lambda)}")
    sheet.cell(row=5, column=7, value=f"{str(cost_lambda_n)}")
    sheet.cell(row=5, column=8, value=f"{str(loop_num)}")
    sheet.cell(row=5, column=9, value=f"{str(decide_val_threshold)}")

    phase_num = sheet.cell(row=10, column=1).value
    row_start = 13
    col_start = 3 * phase_num - 2

    if int(phase) != phase_num - 1:
        logging.error(f"現在のフェーズ数{phase}が、エクセルに記載のあるフェーズ数{phase_num - 1}と一致しません")
        return False

    logging.info("\n\n")
    logging.info("最適化を開始します")
    logging.info(f"引数のファイル名：{file_path}")

    input_liml_file_name = ''
    if phase_num == 1:
        input_liml_file_name = sys.argv[1]
    else:
        input_liml_file_name = "{}_phase_{}.liml".format(str(sys.argv[1])[:-5], phase_num - 1)

    if check_input_excel(sheet) == False:
        print("Input excel data is invalid")
        return False

    with open(input_liml_file_name, 'r', encoding='utf-8') as f:
        root = ET.fromstring(f.read())

    logging.info(f"解析に使用した入力ファイル名：{input_liml_file_name}")

    merged_node_dict = {}
    for solution in root.findall(".//solution"):
        for node in solution.findall("./node"):
            nid = node.attrib.get('nid')
            x = float(node.attrib.get('x', 0))
            y = float(node.attrib.get('y', 0))
            z = float(node.attrib.get('z', 0))
            merged_node_dict[nid] = {'nid': nid, 'x': x, 'y': y, 'z': z}

    elem_list = []
    for solution in root.findall(".//solution"):
        for elset in solution.findall(".//elset"):
            find_elems_in_elset(elset, elem_list)
    
    mat_list = []
    for solution in root.findall(".//solution"):
        find_mats_in_solution(solution, mat_list)

    elem_results_list = []
    for elem in root.findall(".//results/elem"):
        elem_info = {'eid': elem.attrib.get('eid', None)}
        localnodes = []
        for localnode in elem.findall('localnode'):
            localnodes.append(localnode.attrib)
        elem_info['localnodes'] = localnodes
        elem_results_list.append(elem_info)

    for elem in elem_list:
        node_ids = elem.get('nodes', [])
        node_data = [merged_node_dict.get(str(nid), {}) for nid in node_ids]
        elem['node_data'] = node_data

        area = calculate_quadrilateral_area(node_data)
        elem['area'] = area
        points = []
        for vertex in node_data:
            points.append((float(vertex['x']), float(vertex['y'])))
        polygon = Polygon(points)
        elem['polygon'] = polygon
        # center_x = sum(vertex['x'] for vertex in node_data) / 4
        # center_y = sum(vertex['y'] for vertex in node_data) / 4
        # elem['center_x'] = center_x
        # elem['center_y'] = center_y
        # elem['width_x'] = abs(node_data[1]['x'] - node_data[0]['x'])
        # elem['width_y'] = abs(node_data[2]['y'] - node_data[1]['y'])

        thickness = float(elem.get('thickness', 1))  # Assuming a default thickness of 1 if not available
        elem['volume'] = area * thickness

    stress_avg_count = defaultdict(lambda: {'stressxx': 0.0, 'stressyy': 0.0, 'stressxy': 0.0, 'vonmises': 0.0, 'count': 0})
    for elem in elem_results_list:
        eid = elem.get('eid', None)
        if eid is not None:
            for localnode in elem.get('localnodes', []):
                for stress_type in ['stressxx', 'stressyy', 'stressxy', 'vonmises']:
                    stress_value = localnode.get(stress_type, None)
                    if stress_value is not None:
                        try:
                            stress_value = float(stress_value)
                            stress_avg_count[eid][stress_type] += stress_value
                        except ValueError:
                            pass
                stress_avg_count[eid]['count'] += 1

    localnode_elem_list = []
    for eid, data in stress_avg_count.items():
        if data['count'] > 0:
            avg_data = {k: v / data['count'] for k, v in data.items() if k != 'count'}
            avg_data['eid'] = eid
            localnode_elem_list.append(avg_data)

    for elem_info in elem_list:
        material_name = elem_info.get('material')
        if material_name:
            for mat in mat_list:
                if mat['name'] == material_name:
                    elem_info.update(mat)

    merged_elem_list = []
    elem_dict = {elem['eid']: elem for elem in elem_list}
    localnode_elem_dict = {elem['eid']: elem for elem in localnode_elem_list}

    for eid, elem in elem_dict.items():
        if eid in localnode_elem_dict:
            merged_elem = {**elem, **localnode_elem_dict[eid]}
            merged_elem_list.append(merged_elem) 
    
    end_time_1 = time.time()
    elapsed_time_1 = end_time_1 - start_time_1
    sheet.cell(row=row_start, column=col_start, value="Read Input Data")
    sheet.cell(row=row_start, column=col_start + 1, value=f"{str(elapsed_time_1)}")
    row_start += 1

    bRemainAllEdge = False
    optimize_elem_dict = {}
    density_zero_elem_list = []
    x_start = 0.0
    y_start = 0.0
    x_length = 70.0
    y_length = 40.0
    edge_ratio_x = 40
    edge_ratio_y = 20
    x_upper = x_start + x_length - x_length / edge_ratio_x
    x_lower = x_start + x_length / edge_ratio_x
    y_upper = y_start + y_length - y_length / edge_ratio_y
    y_lower = y_start + y_length / edge_ratio_y
    for index, elem in enumerate(merged_elem_list):
        b_need_optimize = True
        eid = elem.get('eid', 0)
        if bRemainAllEdge == True:
            node_data = elem['node_data']
            for node in node_data:
                if node.get('x', None) <= x_lower:
                    b_need_optimize = False
                    break
                if node.get('x', None) >= x_upper:
                    b_need_optimize = False
                    break
                if node.get('y', None) <= y_lower:
                    b_need_optimize = False
                    break
                if node.get('y', None) >= y_upper:
                    b_need_optimize = False
                    break
        if b_need_optimize == True and phase_num > 1:
            row_start_check_finish = 20
            dens_value_old = float(sheet.cell(row=row_start_check_finish + index + 1, column=col_start - 2).value)
            if dens_value_old >= (1.0 - decide_val_threshold - threshold):
                b_need_optimize = False
            if dens_value_old <= (decide_val_threshold + threshold):
                b_need_optimize = False
                density_zero_elem_list.append(str(eid))
                
        if b_need_optimize == True:
            optimize_elem_dict[eid] = elem

    if len(optimize_elem_dict) == 0:
        logging.info("\n\n")
        logging.info("最適化が完了したため処理を終了します")
        return True
            
    start_time_2 = time.time()

    energy_list_for_scale = []
    volume_list_for_scale = []

    first_density = target_density
    # sum_volume = 0.0
    # l_i_dens_part = 0.0
    # for index, (key, elem) in enumerate(optimize_elem_dict.items()):
    #     volume = float(elem.get('volume', 0))
    #     sum_volume += volume
    #     density_now = 0
    #     if phase_num == 1:
    #         density_now = first_density
    #     else:
    #         eid_row = eid + 20
    #         density_now = sheet.cell(row=eid_row, column=col_start - 2).value
    #     l_i_dens_part += (density_now * volume)
    
    energy_part_elem_dict = {}
    ising_index_eid_map = {}
    nInternalid = len(optimize_elem_dict)
    h = defaultdict(int)
    J = defaultdict(int)
    energy_exp = 0.0
    energy_exp_2_zyou = 0.0
    for index, (key, elem) in enumerate(optimize_elem_dict.items()):
        eid = int(key)
        stressxx = elem.get('stressxx', 0)
        stressyy = elem.get('stressyy', 0)
        stressxy = elem.get('stressxy', 0)
        poissonratio = float(elem.get('poissonratio', 0))
        volume = float(elem.get('volume', 0))

        ising_index_eid_map[eid] = index

        density_now = 0
        if phase_num == 1:
            density_now = first_density
        else:
            eid_row = eid + 20
            density_now = sheet.cell(row=eid_row, column=col_start - 2).value

        density_plus_delta = density_now + density_increment
        density_minus_delta = density_now - density_increment

        alpha_value = pow(density_plus_delta, (1 - density_power))
        beta_value = pow(density_minus_delta, (1 - density_power))

        kappa_i = (pow(stressxx, 2.0) - 2.0 * poissonratio * stressxx * stressyy + pow(stressyy, 2.0) + 2.0 * (1.0 + poissonratio) * pow(stressxy, 2.0)) * volume / initial_youngs_modulus
        
        energy_list_for_scale.append(alpha_value * kappa_i)
        energy_list_for_scale.append(beta_value * kappa_i)

        energy_exp += alpha_value * kappa_i * 0.5 + beta_value * kappa_i * 0.5

        alpha_value_2 = pow(density_plus_delta, (2 * (1 - density_power)))
        beta_value_2 = pow(density_minus_delta, (2 * (1 - density_power)))

        energy_exp_2_zyou += alpha_value_2 * pow(kappa_i, 2) * 0.5 + beta_value_2 * pow(kappa_i, 2) * 0.5

        # volume_list_for_scale.append(alpha_value * volume - target_density * sum_volume / nInternalid)
        # volume_list_for_scale.append(alpha_value * volume - target_density * sum_volume / nInternalid)

        volume_list_for_scale.append(volume)
        volume_list_for_scale.append(-1.0 * volume)

    mean_of_energy_list = np.mean(energy_list_for_scale)
    std_of_energy_list = np.std(energy_list_for_scale)

    # bunsan_e = energy_exp_2_zyou - pow(energy_exp, 2)
    # std_of_energy_list = np.sqrt(bunsan_e)

    # standardized_data = (energy_list_for_scale - mean_of_energy_list) / std_of_energy_list

    # mean_of_volume_list = np.mean(volume_list_for_scale)
    std_of_volume_list = np.std(volume_list_for_scale)

    # standardized_data_2 = (volume_list_for_scale - mean_of_volume_list) / std_of_volume_list

    first_index_list = []
    # second_index_list = []
    # bunsan = calc_seiyaku_bunsan(nInternalid)
    for index, (key, elem) in enumerate(optimize_elem_dict.items()):
        eid = int(key)
        stressxx = elem.get('stressxx', 0)
        stressyy = elem.get('stressyy', 0)
        stressxy = elem.get('stressxy', 0)
        poissonratio = float(elem.get('poissonratio', 0))
        volume = float(elem.get('volume', 0))

        ising_index_eid_map[eid] = index

        density_now = 0
        if phase_num == 1:
            density_now = first_density
        else:
            eid_row = eid + 20
            density_now = sheet.cell(row=eid_row, column=col_start - 2).value

        density_plus_delta = density_now + density_increment
        density_minus_delta = density_now - density_increment

        alpha_value = pow(density_plus_delta, (1 - density_power))
        beta_value = pow(density_minus_delta, (1 - density_power))
        k_0 = (alpha_value - beta_value) / 2.0
        kappa_i = (pow(stressxx, 2.0) - 2.0 * poissonratio * stressxx * stressyy + pow(stressyy, 2.0) + 2.0 * (1.0 + poissonratio) * pow(stressxy, 2.0)) * volume / initial_youngs_modulus
        # l_i_part = l_i_dens_part - target_density * sum_volume - mean_of_volume_list * nInternalid
        # l_i_part = density_now * sum_volume - target_density * sum_volume - mean_of_volume_list * nInternalid + target_density * mean_of_volume_list
        # l_i = l_i_part / std_of_volume_list

        scale_lambda = 1.0
        h_first = k_0 * kappa_i / std_of_energy_list / np.sqrt(nInternalid) / 3.0
        # h_second = 2.0 * cost_lambda * l_i * density_increment * volume / std_of_volume_list
        # h[index] = h_first + h_second
        h[index] = h_first
        first_index_list.append(h_first)
        # second_index_list.append(h_second)

        energy_part_elem_dict[eid] = h[index]

        for j_index in range(index + 1, nInternalid):
            list_key = list(optimize_elem_dict.keys())
            volume_j = optimize_elem_dict[list_key[j_index]].get('volume', 0)
            # J[(index,j_index)] = 2.0 * cost_lambda * density_increment * density_increment * volume * volume_j / std_of_volume_list / std_of_volume_list
            # J[(index,j_index)] = 2.0 * cost_lambda * volume * volume_j / np.sqrt(bunsan)
            J[(index,j_index)] = 2.0 * cost_lambda * volume * volume_j / pow(std_of_volume_list, 2) / nInternalid / 9.0
            # print(J[(index,j_index)])

    # print("energy_part_elem_dict")
    # print(energy_part_elem_dict)
    # print("second_index_list\n")
    # print(second_index_list)

    # plt.figure()
    # plt.plot(first_index_list, marker='o')
    # plt.xlabel('Index')
    # plt.ylabel('First Value')
    # plt.title('First Value Data Plot')
    # temp_first_image_path = "first_image.png"
    # plt.savefig(temp_first_image_path)
    # plt.close()
    # img = Image(temp_first_image_path)
    # new_sheet.add_image(img, 'A25')

    # plt.figure()
    # plt.plot(second_index_list, marker='o')
    # plt.xlabel('Index')
    # plt.ylabel('second Value')
    # plt.title('second Value Data Plot')
    # temp_second_image_path = "seccond_image.png"
    # plt.savefig(temp_second_image_path)
    # plt.close()
    # img = Image(temp_second_image_path)
    # new_sheet.add_image(img, 'K25')

    # energy_part_values = list(energy_part_elem_dict.values())
    # mean_value = np.mean(energy_part_values)
    # median_value = np.median(energy_part_values)
    # max_value = np.amax(energy_part_values)
    # cost_scale = max_value
    # logging.info(f"energy_part_valuesの平均値：{mean_value}")
    # logging.info(f"energy_part_valuesの中央値：{median_value}")
    # logging.info(f"energy_part_valuesの最大値：{max_value}")

    ising_index_dict = {}
    bUseOptimization = True

    if bUseOptimization:
        sampler = LeapHybridSampler()
        response = sampler.sample_ising(h, J)

        for sample, E in response.data(fields=['sample','energy']):
            S_minus_1 = [k for k,v in sample.items() if v == -1]
            S_plus_1 = [k for k,v in sample.items() if v == 1]

            for elem in S_minus_1:
                ising_index_dict[elem] = -1

            for elem in S_plus_1:
                ising_index_dict[elem] = 1

            print(f"イジングモデルの各要素の最適化後の値は: {ising_index_dict} となる")
    else:
        sorted_energy_part_elem_dict = sorted(energy_part_elem_dict.items(), key=lambda x: x[1], reverse=True)
        half_size = len(sorted_energy_part_elem_dict) // 2

        temp_ising_index_dict = {}
        for i, (key, value) in enumerate(sorted_energy_part_elem_dict):
            isingId = key - 1
            if i < half_size:
                temp_ising_index_dict[isingId] = -1
            else:
                temp_ising_index_dict[isingId] = 1

        ising_index_dict = dict(sorted(temp_ising_index_dict.items()))

    start_time_3 = time.time()
    mat_youngmodulus = {}
    row_start = 19
    sheet.cell(row=row_start, column=col_start, value=f"phase_{phase_num}")
    row_start += 1
    sheet.cell(row=row_start, column=col_start, value="Element number")
    sheet.cell(row=row_start, column=col_start + 1, value="Density")
    sheet.cell(row=row_start, column=col_start + 2, value="Energy part")
    width = initial_condition_data["width"]
    height = initial_condition_data["height"]
    div_x = initial_condition_data['devide_num_x']
    div_y = initial_condition_data['devide_num_y']
    data = np.zeros((div_y, div_x))
    # data2 = np.zeros((div_y, div_x))

    for index, value in enumerate(merged_elem_list):
        eid = index + 1
        sheet.cell(row=row_start + index + 1, column=col_start, value=eid)
        dens_value = 0
        if str(eid) not in optimize_elem_dict:
            if str(eid) in density_zero_elem_list:
                dens_value = 1.0e-9
            else:
                dens_value = 1.0
        else:
            dens_value_old = first_density
            if not phase_num == 1:
                dens_value_old = float(sheet.cell(row=row_start + index + 1, column=col_start - 2).value)
            ising_index = ising_index_eid_map[eid]
            ising_value = ising_index_dict[ising_index] if bUseOptimization else ising_index_dict[index]
            # ising_value = ising_index_dict[ising_index]
            dens_value = dens_value_old + density_increment * ising_value
            # if dens_value >= (1.0 - decide_val_threshold - threshold):
            #     dens_value = 1.0
            # if dens_value <= (decide_val_threshold + threshold):
            #     dens_value = 0.00001

        sheet.cell(row=row_start + index + 1, column=col_start + 1, value=float(dens_value))

        energy_part = energy_part_elem_dict.get(eid, 0)
        sheet.cell(row=row_start + index + 1, column=col_start + 2, value=float(energy_part))

        mat_youngmodulus[str(eid)] = pow(dens_value, density_power) * initial_youngs_modulus

        # center_x = next((elem['center_x'] for elem in merged_elem_list if str(elem['eid']) == str(eid)), None)
        # center_y = next((elem['center_y'] for elem in merged_elem_list if str(elem['eid']) == str(eid)), None)
        # cell_width = next((elem['width_x'] for elem in merged_elem_list if str(elem['eid']) == str(eid)), None)
        # cell_height = next((elem['width_y'] for elem in merged_elem_list if str(elem['eid']) == str(eid)), None)
        # cell_x = int(center_x // cell_width)
        # cell_y = int(center_y // cell_height)
        # data[cell_y, cell_x] = dens_value
        # data2[cell_y, cell_x] = energy_part

        range_polygon = next((elem['polygon'] for elem in merged_elem_list if str(elem['eid']) == str(eid)), None)
        block_width, block_height = width / div_x, height / div_y
        min_x, min_y, max_x, max_y = range_polygon.bounds
        start_x = max(int(min_x // block_width), 0)
        end_x = min(int(np.ceil(max_x / block_width)), div_x)
        start_y = max(int(min_y // block_height), 0)
        end_y = min(int(np.ceil(max_y / block_height)), div_y)
        for i in range(start_x, end_x):
            for j in range(start_y, end_y):
                block_right_top = Point((i + 1) * block_width, (j + 1) * block_height)
                if range_polygon.contains(block_right_top):
                    data[j, i] = dens_value

    plt.imshow(data, cmap='gray_r', origin='lower', extent=[0, width, 0, height], vmin=0, vmax=1.1)
    plt.colorbar()
    plt.title("Density distribution of elements")
    plt.xlabel("x")
    plt.ylabel("y")
    temp_image_path = "optimize_cae_density_temp.png"
    plt.savefig(temp_image_path)
    plt.close()

    # plt.imshow(data2, cmap='viridis', origin='lower', extent=[0, width, 0, height])
    # plt.colorbar()
    # plt.title("distribution of {νσi}T{σi}vi/E0")
    # plt.xlabel("x")
    # plt.ylabel("y")
    # temp_image_path_2 = "optimize_cae_density_temp_2.png"
    # plt.savefig(temp_image_path_2)
    # plt.close()

    new_sheet_name = "Image on phase " + str(phase_num)
    new_sheet = workbook.create_sheet(new_sheet_name)

    new_sheet['A1'] = '要素の密度分布'
    img = Image(temp_image_path)
    new_sheet.add_image(img, 'A3')

    # new_sheet['K1'] = '{νσi}T{σi}vi/E0 の分布'
    # img = Image(temp_image_path_2)
    # new_sheet.add_image(img, 'K3')

    phase_num += 1
    sheet.cell(row=10, column=1, value=phase_num)

    row_start = 15
    end_time_3 = time.time()
    elapsed_time_3 = end_time_3 - start_time_3
    sheet.cell(row=row_start, column=col_start, value="Update excel file")
    sheet.cell(row=row_start, column=col_start + 1, value=f"{str(elapsed_time_3)}")
    row_start += 1

    start_time_4 = time.time()
    for solution in root.findall(".//solution"):
        for child in list(solution):
            solution.remove(child)

        new_analysis = ET.Element('analysis', type="S30")
        new_elset = ET.Element('elset', name="Default", color="-6710887")
        new_table = ET.Element('table')

        solution.append(new_analysis)
        solution.append(new_elset)
        solution.append(new_table)

    elsets_to_modify = []

    for elset in root.findall('.//elset'):
        elems = elset.findall('elem')
        if len(elems) > 1:
            elsets_to_modify.append(elset)

    for elset in elsets_to_modify:
        elems = elset.findall('elem')
        for elem in elems:
            eid = elem.get('eid', 0)
            new_elset = ET.Element('elset')
            for attr_name, attr_value in elset.attrib.items():
                new_attr_value = 0
                if attr_name == "name":
                    new_attr_value = "Component(" + str(eid) + ")"
                if attr_name == "material":
                    new_attr_value = "Material(" + str(eid) + ")"
                if attr_name == "color":
                    new_attr_value = "-8209549"
                new_elset.set(attr_name, new_attr_value)

            new_elset.append(elem) 
            root.append(new_elset)
            mat_tags = root.findall(".//mat[@name='Material']")
            for mat in mat_tags:
                new_mat = ET.Element('mat', mat.attrib)
                new_mat.set('name', f'Material({eid})')
                for mat_child in mat:
                    if mat_child.tag == 'geometric':
                        new_elem = ET.Element('geometric', mat_child.attrib)
                    if mat_child.tag == 'mechanical':
                        new_elem = ET.Element('mechanical', mat_child.attrib)                  
                    new_mat.append(new_elem)
                root.append(new_mat)
        for elem in elems:
            elset.remove(elem)

    for mat in root.findall(".//mat"):
        mat_name = mat.attrib.get('name', None)
        match = re.search(r'\(\d+\)', mat_name)
        if not match:
            continue
        eid_split = mat_name.split("(")[1].split(")")[0]
        youngmodulus_value = mat_youngmodulus.get(eid_split, None)
        if youngmodulus_value is None:
            continue
        mechanical_tag = mat.find("./mechanical")
        if mechanical_tag is not None:
            mechanical_tag.set("youngsmodulus", str(youngmodulus_value))
    
    # bDoSolver = True

    # if bDoSolver:
    new_file_name = increment_phase_number(input_liml_file_name)
    new_file_content = ET.tostring(root, encoding='unicode')
    with open(new_file_name, 'w', encoding='utf-8') as f:
        f.write(new_file_content)

    logging.info(f"最適化後のファイル名：{new_file_name}")

    hwndvscode = find_vscode_window()
    if hwndvscode:
        win32gui.ShowWindow(hwndvscode, win32con.SW_MINIMIZE)
        # win32gui.SetWindowPos(hwndvscode, win32con.HWND_BOTTOM, 0, 0, 0, 0, win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)
    proc = subprocess.run(["explorer", new_file_name])
    time.sleep(15)
    hwnd = win32gui.FindWindow(None, f"{new_file_name} - LISA")
    if hwnd:
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
        win32gui.SetForegroundWindow(hwnd)
    pyautogui.moveTo(158, 70, duration=1)
    pyautogui.click()
    time.sleep(25)
    pyautogui.moveTo(1886, 4)
    pyautogui.click()
    time.sleep(3)
    pyautogui.moveTo(867, 594)
    pyautogui.click()

    end_time_4 = time.time()
    elapsed_time_4 = end_time_4 - start_time_4
    sheet.cell(row=row_start, column=col_start, value="Put LISA file and do FEM solver")
    sheet.cell(row=row_start, column=col_start + 1, value=f"{str(elapsed_time_4)}")

    workbook.save(sys.argv[2])
    os.remove(temp_image_path)
    # os.remove(temp_image_path_2)
    # os.remove(temp_first_image_path)
    # os.remove(temp_second_image_path)

    print(f"success optimization on phase {phase_num - 1}")

    return True

if __name__ == '__main__':
    sys.argv = ["cae_optimize.py", "C:\\work\\github\\q-annealing-d-wave-test\\cantilever_different_volume_1.liml", "C:\\work\\github\\q-annealing-d-wave-test\\result_summary.xlsx"]
    if len(sys.argv) < 3:
        print("Usage: python merged_cae_test.py <liml_file_path> <excel_file_path>")
    else:
        main(sys.argv[1])
